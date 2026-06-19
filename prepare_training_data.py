# -*- coding: utf-8 -*-
"""
Связка факт-часов из 1С с признаками из спецификаций -> датасет для обучения.

Берёт выгрузку 1С (Шаблон_выгрузки_1С.xlsx) и папку со спецификациями, для каждого
проекта парсит признаки (parse_spec) и присоединяет фактические нормо-часы. Результат —
CSV в том же формате, что и synthetic_dataset.csv, готовый для train_model.py.

Ключ соединения: столбец «Файл спецификации» в выгрузке = имя файла в папке --specs-dir.

Запуск:
    python3 prepare_training_data.py --hours Шаблон_выгрузки_1С.xlsx \
            --specs-dir specs/ --out training_data.csv

Зависимости: openpyxl, pandas (+ pdfplumber для PDF-спецификаций).
"""
import os
import argparse
import pandas as pd

import parse_spec
from norms_model import ALL_FEATURES, estimate_hours

# Ключи столбцов выгрузки 1С (распознаются по подстроке, регистр не важен)
COL = {
    "file":    ["файл специф", "файл", "спецификац"],
    "total":   ["итого", "всего часов", "трудоёмкость"],
    "outlier": ["выброс", "исключ", "брак"],
    "name":    ["обознач", "чертёж", "шкаф"],
}
# столбцы часов по этапам — для фоллбэка, если ИТОГО задано формулой и не закэшировано
STAGE_KEYS = ["мехобраб", "силов", "слаботоч", "пнр", "наладк"]


def _to_float(v):
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, AttributeError):
        return None


def _find(headers, role):
    for j, h in enumerate(headers):
        hl = str(h).strip().lower()
        if any(k in hl for k in COL[role]):
            return j
    return None


def read_hours(path, sheet=None):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    grid = [["" if c is None else c for c in r] for r in ws.iter_rows(values_only=True)]
    grid = [r for r in grid if any(str(c).strip() for c in r)]
    # строка заголовка = где находится «файл специф.» и «итого»
    hi = 0
    for i, r in enumerate(grid[:25]):
        if _find(r, "file") is not None and _find(r, "total") is not None:
            hi = i; break
    hdr = grid[hi]
    ci = {role: _find(hdr, role) for role in COL}
    if ci["file"] is None:
        raise SystemExit("В выгрузке не найден столбец «Файл спецификации».")
    stage_cols = [j for j, h in enumerate(hdr)
                  if any(k in str(h).strip().lower() for k in STAGE_KEYS)]
    out = []
    for r in grid[hi + 1:]:
        get = lambda j: (r[j] if (j is not None and j < len(r)) else "")
        f = str(get(ci["file"])).strip()
        if not f:
            continue
        total = _to_float(get(ci["total"])) if ci["total"] is not None else None
        if total is None:   # ИТОГО пустое/формула -> сумма по этапам
            parts = [_to_float(get(j)) for j in stage_cols]
            parts = [p for p in parts if p is not None]
            total = sum(parts) if parts else None
        out.append({"file": f, "total": total,
                    "outlier": str(get(ci["outlier"])).strip(),
                    "name": str(get(ci["name"])).strip()})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--hours", required=True, help="выгрузка 1С (.xlsx)")
    ap.add_argument("--specs-dir", required=True, help="папка со спецификациями")
    ap.add_argument("--out", default="training_data.csv")
    ap.add_argument("--sheet", default=None)
    args = ap.parse_args()

    records = read_hours(args.hours, args.sheet)
    rows, skipped, missing = [], 0, []
    for rec in records:
        if rec["outlier"] in ("1", "да", "true", "x", "+"):
            skipped += 1
            continue
        path = os.path.join(args.specs_dir, rec["file"])
        if not os.path.exists(path):
            missing.append(rec["file"]); continue
        try:
            feats, _ = parse_spec.build_features(path)
        except Exception as e:
            missing.append(f"{rec['file']} (ошибка: {e})"); continue
        actual = rec["total"]
        if actual is None:
            missing.append(f"{rec['file']} (нет числа в ИТОГО)"); continue
        row = {k: feats.get(k) for k in ALL_FEATURES}
        for k in ("mount_panel", "din_inclined"):
            row[k] = int(bool(row.get(k)))
        row["baseline_hours"] = round(estimate_hours(feats), 2)
        row["actual_hours"] = actual
        rows.append(row)

    if not rows:
        raise SystemExit("Не собрано ни одной строки — проверьте имена файлов и папку --specs-dir.")
    df = pd.DataFrame(rows)[ALL_FEATURES + ["baseline_hours", "actual_hours"]]
    df.to_csv(args.out, index=False, encoding="utf-8-sig")

    print(f"Готово: {args.out}  ({len(df)} проектов)")
    print(f"  пропущено как выбросы: {skipped}")
    if missing:
        print(f"  не найдено/ошибка ({len(missing)}): " + ", ".join(missing[:8])
              + (" …" if len(missing) > 8 else ""))
    mape = (abs(df.actual_hours - df.baseline_hours) / df.actual_hours).mean() * 100
    print(f"  MAPE параметрической модели на реальных данных: {mape:.1f}%")
    print(f"\nДалее:  python3 train_model.py --data {args.out}")


if __name__ == "__main__":
    main()
