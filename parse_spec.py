#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Парсер спецификации ШАУ -> вектор признаков для модели нормирования трудоёмкости.

Поддерживаемые форматы входа (определяются по расширению):
  * .pdf            — PDF-экспорт EPLAN/AutoCAD Electrical/аналог (по тексту);
  * .xlsx / .xls    — выгрузка перечня изделий (Parts list / Stückliste) EPLAN, Excel;
  * .csv            — табличная выгрузка BOM.

Назначение:
  1) Извлечь BOM (Спецификацию), параметры корпуса, число вводов, клеммы, аналог. цепи;
  2) Свести их в плоский вектор признаков (features) согласно разделу 2 ТЗ;
  3) Этот же вектор подаётся в калькулятор (norms_model) и в ML-модель (train_model).

Использование:
    python3 parse_spec.py "spec.pdf"
    python3 parse_spec.py "bom.xlsx" --json out.json --bom
    python3 parse_spec.py "bom.csv"  --sheet "Перечень элементов"

Зависимости: pdfplumber (для PDF), openpyxl/pandas (для Excel/CSV)
"""

import os
import re
import json
import argparse


# --------------------------------------------------------------------------
# Словари ключевых слов: по наименованию позиции BOM определяем категорию.
# Список расширяемый — под номенклатуру конкретного предприятия.
# --------------------------------------------------------------------------
KEYWORDS = {
    "freq_converter":   ["частотн", "преобразователь частоты", "nci-", "instart nci"],
    "soft_starter":     ["плавн", "упп ", "soft start"],
    "breaker_le63":     ["автоматический выключатель", "авт. выкл", "ва-101", "ва103", "ва-103"],
    "switch_load":      ["выключатель нагрузки", "рубильник", "nf2-"],
    "contactor":        ["контактор", "пускатель", "кми", "lc1"],
    "relay":            ["реле rf", "rft2co", "промежуточное реле"],
    "phase_relay":      ["реле контроля фаз", "ел-11"],
    "rcd":              ["узо", "дифавтомат", "диф. авт", "дифференц"],
    "plc":              ["программируемое реле", "плк", "пр103", "пр200", "контроллер"],
    "hmi":              ["операторская панель", "панель оператора", "hmi", "rsc-7", "сенсорн"],
    "power_supply":     ["блок питания", "mdr-", "drp-", "mean well"],
    "signal_lamp":      ["лампа сигнальная", "лампа сигн", "индикатор"],
    "fan":              ["вентилятор", "kipvent", "решётк", "решетк"],
    "thermostat":       ["термостат", "термореле"],
    "fuse_holder":      ["держатель предохранителя", "ask 2", "ask2"],
    "terminal":         ["клемма", "клеммник"],
    "jumper":           ["перемычк"],
}


def _to_int(s):
    try:
        return int(re.sub(r"[^\d]", "", s))
    except Exception:
        return 0


def parse_bom(pages_text):
    """Извлекает позиции BOM из ТЕКСТА PDF: [{'name':..., 'qty':int}, ...].

    Строка спецификации имеет жёсткий формат:
        <Поз> <Обозначение> <Наименование...> <Кол-во> <Заказной №> <Изготовитель>
    Принимаем только строки, которые начинаются с номера позиции и заканчиваются
    "...<кол-во> <заказной№> <изготовитель>". Это отсекает текст со схемных листов.
    """
    rows = []
    head_re = re.compile(r"^\d{1,2}\s+\S")              # начинается с № позиции
    qty_re = re.compile(r"\s(\d{1,3})\s+\S+\s+\S+\s*$")  # ...кол-во заказной№ изготовитель
    for txt in pages_text:
        for raw in txt.splitlines():
            line = raw.strip()
            if not line or not head_re.match(line):
                continue
            low = line.lower()
            if not any(k in low for v in KEYWORDS.values() for k in v):
                continue
            m = qty_re.search(line)
            qty = _to_int(m.group(1)) if m else 1
            rows.append({"name": line, "qty": qty or 1})
    return rows


# --------------------------------------------------------------------------
# Загрузчики BOM из табличных форматов (EPLAN Parts list / Excel / CSV).
# Заголовки распознаются по словарю на RU/EN/DE — под разные локали EPLAN.
# --------------------------------------------------------------------------
COL_KEYS = {
    "qty":   ["кол", "колич", "qty", "quantity", "menge", "anzahl", "шт"],
    "name":  ["наимен", "описан", "designation", "description", "bezeichnung", "benennung"],
    "desig": ["обознач", "позиц", "dt", "betriebsmittel", "marking", "device tag", "ozn"],
    "maker": ["изготов", "производ", "manufacturer", "hersteller", "поставщик"],
    "part":  ["заказн", "артикул", "part number", "type number", "artikel", "номер модели"],
}


def _match_col(header, role):
    h = str(header).strip().lower()
    return any(k in h for k in COL_KEYS[role])


def _find_header_row(grid, max_scan=20):
    """Находит строку заголовка = строка с максимумом совпадений по ключам столбцов."""
    best_i, best_score = 0, -1
    for i, row in enumerate(grid[:max_scan]):
        score = sum(any(_match_col(c, role) for c in row) for role in COL_KEYS)
        if score > best_score:
            best_i, best_score = i, score
    return best_i if best_score >= 2 else 0


def _map_columns(header):
    """role -> индекс столбца."""
    cols = {}
    for role in COL_KEYS:
        for j, h in enumerate(header):
            if _match_col(h, role):
                cols[role] = j
                break
    return cols


def load_bom_table(path, sheet=None):
    """Грузит BOM из Excel/CSV -> (rows, full_text).

    rows = [{'name','qty'}], где name = «обозначение + наименование», qty — из столбца
    количества (надёжнее, чем парсинг текста). full_text — все ячейки для разбора
    корпуса/IP/интерфейсов.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        import csv
        with open(path, encoding="utf-8-sig", newline="") as f:
            grid = [list(r) for r in csv.reader(f)]
    else:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        grid = [["" if c is None else c for c in row]
                for row in ws.iter_rows(values_only=True)]
    grid = [r for r in grid if any(str(c).strip() for c in r)]
    if not grid:
        return [], ""

    hi = _find_header_row(grid)
    header = [str(c).strip() for c in grid[hi]]
    cols = _map_columns(header)
    rows = []
    for raw in grid[hi + 1:]:
        cell = lambda role: (str(raw[cols[role]]).strip()
                             if role in cols and cols[role] < len(raw) else "")
        name = (cell("desig") + " " + cell("name")).strip()
        if not name:
            continue
        qty = _to_int(cell("qty")) or 1
        rows.append({"name": name, "qty": qty})
    full_text = "\n".join(" ".join(str(c) for c in r) for r in grid)
    return rows, full_text


# Аксессуары, которые не считаем самостоятельными аппаратами
ACCESSORY = ["цоколь", "скоба", "держатель шильдик", "изолятор", "торцев", "маркир"]


def categorize(rows):
    counts = {k: 0 for k in KEYWORDS}
    for r in rows:
        low = r["name"].lower()
        r["matched"] = False
        is_accessory = any(a in low for a in ACCESSORY)
        for cat, kws in KEYWORDS.items():
            # аксессуары не считаем как аппараты (реле/контакторы/...), но клеммы считаем
            if is_accessory and cat not in ("terminal", "fuse_holder", "jumper"):
                continue
            if any(k in low for k in kws):
                counts[cat] += r["qty"]
                r["matched"] = True
                r["cat"] = cat
                break
    return counts


def parse_corpus(full_text):
    res = {}
    ip = re.search(r"\bIP\s?(21|54|65|66)\b", full_text, re.I)
    res["ip"] = "IP" + ip.group(1) if ip else None
    dim = re.search(r"(\d{3,4})\s?[xх]\s?(\d{3,4})\s?[xх]\s?(\d{2,4})\s?мм", full_text, re.I)
    if dim:
        res["w"], res["h"], res["d"] = (int(dim.group(i)) for i in (1, 2, 3))
    res["mount_panel"] = bool(re.search(r"монтажн\w* панел", full_text, re.I))
    res["din_inclined"] = bool(re.search(r"наклонн\w*\s+DIN", full_text, re.I))
    return res


def parse_holes(full_text):
    total, detail = 0, []
    for m in re.finditer(r"(\d{1,3})\s*отверст\w*\s*[ØΦ⌀Oo]?\s*(\d{1,3})\s*мм", full_text, re.I):
        n, d = int(m.group(1)), int(m.group(2))
        total += n
        detail.append({"count": n, "dia_mm": d})
    return total, detail


def parse_analog(full_text):
    n_420 = len(re.findall(r"4\s*-\s*20\s*мА", full_text, re.I)) \
        + len(re.findall(r"4-20\s*m[aА]", full_text, re.I))
    n_010 = len(re.findall(r"0\s*-\s*10\s*[ВV]", full_text, re.I))
    return n_420, n_010


def parse_interfaces(full_text):
    return {
        "ethernet": bool(re.search(r"ethernet|rj45", full_text, re.I)),
        "rs485": bool(re.search(r"rs\s?-?\s?485|modbus", full_text, re.I)),
        "modbus": bool(re.search(r"modbus", full_text, re.I)),
        "profibus": bool(re.search(r"profibus", full_text, re.I)),
    }


def load_source(path, sheet=None):
    """Диспетчер по расширению -> (rows, full_text, source_kind)."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        try:
            import pdfplumber
        except ImportError:
            raise RuntimeError("Для PDF нужен пакет pdfplumber: pip install pdfplumber")
        with pdfplumber.open(path) as pdf:
            pages_text = [(p.extract_text() or "") for p in pdf.pages]
        full = "\n".join(pages_text)
        return parse_bom(pages_text), full, "pdf"
    if ext in (".xlsx", ".xls", ".xlsm", ".csv"):
        try:
            rows, full = load_bom_table(path, sheet)
        except ImportError:
            raise RuntimeError("Для Excel нужен пакет openpyxl: pip install openpyxl")
        return rows, full, "table"
    raise RuntimeError(f"Неподдерживаемый формат: {ext or '(без расширения)'}. "
                       f"Поддерживаются .pdf, .xlsx, .xls, .csv")


def build_features(path, sheet=None):
    rows, full, kind = load_source(path, sheet)
    counts = categorize(rows)
    corpus = parse_corpus(full)
    holes_total, holes_detail = parse_holes(full)
    n_420, n_010 = parse_analog(full)
    ifaces = parse_interfaces(full)

    features = {
        "ip_rating": corpus.get("ip"),
        "width_mm": corpus.get("w"),
        "height_mm": corpus.get("h"),
        "depth_mm": corpus.get("d"),
        "mount_panel": corpus.get("mount_panel"),
        "din_inclined": corpus.get("din_inclined"),
        "cable_entries": holes_total,
        "cable_entries_detail": holes_detail,
        "breakers_le63": counts["breaker_le63"],
        "load_switches": counts["switch_load"],
        "contactors": counts["contactor"],
        "rcd": counts["rcd"],
        "freq_converters": counts["freq_converter"],
        "soft_starters": counts["soft_starter"],
        "plc": counts["plc"],
        "hmi": counts["hmi"],
        "power_supplies": counts["power_supply"],
        "signal_lamps": counts["signal_lamp"],
        "relays": counts["relay"],
        "phase_relays": counts["phase_relay"],
        "analog_4_20mA": n_420,
        "analog_0_10V": n_010,
        "interfaces": ifaces,
        "interfaces_count": sum(1 for v in ifaces.values() if v),
        "terminals": counts["terminal"],
        "fuse_holders": counts["fuse_holder"],
        "fans": counts["fan"],
        "thermostats": counts["thermostat"],
        "_source": kind,
    }
    return features, rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("source", help="спецификация: .pdf / .xlsx / .csv")
    ap.add_argument("--json")
    ap.add_argument("--sheet", help="имя листа Excel (если не первый)")
    ap.add_argument("--bom", action="store_true")
    args = ap.parse_args()

    features, rows = build_features(args.source, args.sheet)
    if features.get("_source") == "table" and not features["cable_entries"]:
        print("[i] Источник — табличный BOM: число вводов/аналог. цепей со схемы "
              "недоступно (берётся со схемы расположения/клемм в PDF).\n")

    print("=== ВЕКТОР ПРИЗНАКОВ ===")
    for k, v in features.items():
        if k in ("cable_entries_detail", "_source"):
            continue
        print(f"{k:22}: {v}")
    if features["cable_entries_detail"]:
        print("  отверстия:", features["cable_entries_detail"])

    if args.bom:
        print("\n=== Распознанные позиции BOM ===")
        for r in rows:
            flag = "" if r["matched"] else "  [не классиф.]"
            print(f"  x{r['qty']:>3}  {r['name'][:78]}{flag}")

    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(features, f, ensure_ascii=False, indent=2)
        print(f"\nСохранено: {args.json}")


if __name__ == "__main__":
    main()
